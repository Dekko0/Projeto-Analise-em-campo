import streamlit as st
import pandas as pd
import io
import json
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
from datetime import datetime

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Levantamento de Cargas", layout="wide", page_icon="⚡")

PLANILHA_PADRAO = "Levantamento_Base.xlsx"
USUARIOS_FILE = "usuarios.json"

# --- FUNÇÕES DE GESTÃO DE USUÁRIOS ---

def carregar_usuarios():
    if os.path.exists(USUARIOS_FILE):
        with open(USUARIOS_FILE, "r") as f:
            return json.load(f)
    return {"Admin": "cargas2024"}

def salvar_usuarios(usuarios):
    with open(USUARIOS_FILE, "w") as f:
        json.dump(usuarios, f)

# --- FUNÇÕES DE PERSISTÊNCIA ---

def get_user_db_path(nome_usuario=None):
    user = nome_usuario if nome_usuario else st.session_state.get('usuario_ativo')
    if user:
        nome_limpo = "".join(filter(str.isalnum, user))
        return f"dados_{nome_limpo}.json"
    return None

def salvar_dados_locais(dados):
    path = get_user_db_path()
    if path:
        with open(path, "w") as f:
            json.dump(dados, f)

def carregar_dados_locais(path_especifico=None):
    path = path_especifico if path_especifico else get_user_db_path()
    if path and os.path.exists(path):
        with open(path, "r") as f:
            return json.load(f)
    return []

# --- MODAIS DE CONFIRMAÇÃO ---

@st.dialog("Confirmar Exclusão")
def confirmar_exclusao_dialog(index=None, tipo="individual"):
    usuarios = carregar_usuarios()
    senha_correta = usuarios.get(st.session_state['usuario_ativo'])
    
    if tipo == "individual":
        item = st.session_state['db_formularios'][index]
        st.warning(f"Excluir levantamento UC: {item['cod_instalacao']}?")
    else:
        st.error("⚠️ ATENÇÃO: Isso apagará TODOS os seus levantamentos salvos!")

    senha_confirmacao = st.text_input("Sua senha de usuário", type="password", key="confirm_pass_local")
    if st.button("Confirmar Exclusão", type="primary", use_container_width=True):
        if senha_confirmacao == senha_correta:
            if tipo == "individual":
                st.session_state['db_formularios'].pop(index)
            else:
                st.session_state['db_formularios'] = []
            salvar_dados_locais(st.session_state['db_formularios'])
            st.rerun()
        else:
            st.error("Senha incorreta!")

@st.dialog("Remover Usuário")
def excluir_usuario_dialog(nome_usuario):
    st.warning(f"Remover acesso de **{nome_usuario}**?")
    usuarios = carregar_usuarios()
    senha_admin = st.text_input("Senha do Admin", type="password", key="confirm_pass_admin_user")
    
    if st.button("Confirmar Remoção de Acesso", type="primary", use_container_width=True):
        if senha_admin == usuarios.get("Admin"):
            del usuarios[nome_usuario]
            salvar_usuarios(usuarios)
            st.success("Acesso removido!")
            st.rerun()
        else: st.error("Senha incorreta!")

@st.dialog("Exclusão Permanente de Arquivo")
def excluir_arquivo_permanente_dialog(caminho_arquivo):
    st.error(f"⚠️ ATENÇÃO CRÍTICA: Você está apagando o arquivo físico: **{caminho_arquivo}**")
    st.warning("Esta ação é irreversível. Todos os dados contidos neste arquivo serão deletados do servidor.")
    
    usuarios = carregar_usuarios()
    senha_admin = st.text_input("Senha do Administrador para autorizar", type="password", key="confirm_pass_admin_file")
    
    if st.button("EXCLUIR PERMANENTEMENTE", type="primary", use_container_width=True):
        if senha_admin == usuarios.get("Admin"):
            try:
                os.remove(caminho_arquivo)
                st.success("Arquivo deletado com sucesso!")
                st.rerun()
            except Exception as e:
                st.error(f"Erro ao deletar: {e}")
        else:
            st.error("Senha de administrador incorreta!")

# --- FUNÇÕES DE EXCEL E EMAIL ---

def exportar_para_excel(dados_lista):
    if 'planilha_modelo' not in st.session_state or not st.session_state['planilha_modelo']:
        return None
    st.session_state['planilha_modelo'].seek(0)
    book = load_workbook(st.session_state['planilha_modelo'])
    for registro in dados_lista:
        if registro['tipo_equipamento'] in book.sheetnames:
            sheet = book[registro['tipo_equipamento']]
            colunas_excel = [cell.value for cell in sheet[1]]
            nova_linha = [registro['dados'].get(col, "") for col in colunas_excel]
            sheet.append(nova_linha)
    output = io.BytesIO()
    book.save(output)
    output.seek(0)
    return output

def enviar_email(arquivo_buffer, destinatario):
    try:
        msg = MIMEMultipart()
        msg['From'] = "levantamento.poupenergia@gmail.com"
        msg['To'] = destinatario
        msg['Subject'] = f"Levantamento {st.session_state['usuario_ativo']} - {datetime.now().strftime('%d/%m/%Y')}"
        msg.attach(MIMEText("Planilha em anexo.", 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(arquivo_buffer.getvalue())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename=levantamento.xlsx')
        msg.attach(part)
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("levantamento.poupenergia@gmail.com", "kiqplowxqprcugjc")
        server.send_message(msg)
        server.quit()
        return True
    except: return False

def analisar_modelo_excel(file_content):
    buffer = io.BytesIO(file_content) if isinstance(file_content, bytes) else io.BytesIO(file_content.getvalue())
    wb = load_workbook(buffer, data_only=True)
    estrutura = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        headers = []
        for cell in sheet[1]:
            if cell.value: headers.append({"nome": str(cell.value), "col_letter": cell.column_letter, "tipo": "texto", "opcoes": []})
        for dv in sheet.data_validations.dataValidation:
            if dv.type == "list":
                for ref in str(dv.sqref).split():
                    col_letter = "".join(filter(str.isalpha, ref.split(':')[0]))
                    for h in headers:
                        if h["col_letter"] == col_letter:
                            h["tipo"] = "selecao"
                            formula = dv.formula1
                            if formula and formula.startswith('"'): h["opcoes"] = formula.strip('"').split(',')
        estrutura[sheet_name] = [{"nome": h["nome"], "tipo": h["tipo"], "opcoes": h["opcoes"]} for h in headers]
    return estrutura

# --- INICIALIZAÇÃO ---

if 'usuario_ativo' not in st.session_state: st.session_state['usuario_ativo'] = None
if 'form_id' not in st.session_state: st.session_state['form_id'] = 0

if not st.session_state['usuario_ativo']:
    st.markdown("<h1 style='text-align: center;'>⚡ Sistema de Cargas</h1>", unsafe_allow_html=True)
    usuarios_db = carregar_usuarios()
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        u = st.selectbox("Técnico", options=["Selecione..."] + list(usuarios_db.keys()))
        p = st.text_input("Senha", type="password")
        if st.button("Entrar", use_container_width=True):
            if u in usuarios_db and p == usuarios_db[u]:
                st.session_state['usuario_ativo'] = u
                st.session_state['db_formularios'] = carregar_dados_locais()
                st.rerun()
            else: st.error("Incorreto.")
    st.stop()

if 'planilha_modelo' not in st.session_state:
    if os.path.exists(PLANILHA_PADRAO):
        with open(PLANILHA_PADRAO, "rb") as f:
            content = f.read()
            st.session_state['planilha_modelo'] = io.BytesIO(content)
            st.session_state['estrutura_modelo'] = analisar_modelo_excel(content)

# --- INTERFACE ---

st.sidebar.title(f"👤 {st.session_state['usuario_ativo']}")
if st.sidebar.button("Sair"):
    st.session_state['usuario_ativo'] = None
    st.rerun()

menu = st.sidebar.radio("Navegação", ["1. Preenchimento", "2. Exportar & Listar"] + (["3. Painel Admin"] if st.session_state['usuario_ativo'] == "Admin" else []))

if menu == "1. Preenchimento":
    st.header("📝 Novo Levantamento")
    if 'estrutura_modelo' in st.session_state:
        tipo = st.selectbox("Tipo de Equipamento", options=list(st.session_state['estrutura_modelo'].keys()), 
                            on_change=lambda: st.session_state.update({"form_id": st.session_state['form_id']+1}))
        uc = st.text_input("UC", key=f"uc_{st.session_state['form_id']}")
        campos = st.session_state['estrutura_modelo'][tipo]
        respostas = {}
        with st.form(key=f"f_{st.session_state['form_id']}"):
            cols = st.columns(2)
            for i, c in enumerate(campos):
                col_at = cols[i % 2]
                if c['tipo'] == 'selecao': respostas[c['nome']] = col_at.selectbox(c['nome'], options=c['opcoes'], key=f"in_{st.session_state['form_id']}_{c['nome']}")
                else: respostas[c['nome']] = col_at.text_input(c['nome'], key=f"in_{st.session_state['form_id']}_{c['nome']}")
            if st.form_submit_button("➕ Salvar"):
                if uc:
                    novo = {"cod_instalacao": uc, "tipo_equipamento": tipo, "data_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"), "dados": respostas}
                    st.session_state['db_formularios'].append(novo)
                    salvar_dados_locais(st.session_state['db_formularios'])
                    st.session_state['form_id'] += 1
                    st.rerun()

elif menu == "2. Exportar & Listar":
    st.header("📋 Seus Levantamentos")
    if not st.session_state['db_formularios']: 
        st.info("Sem dados salvos localmente.")
    else:
        for idx, item in enumerate(st.session_state['db_formularios']):
            with st.container(border=True):
                c1, c2, c3, c4 = st.columns([2, 2, 2, 1])
                c1.write(f"**UC:** {item['cod_instalacao']}")
                c2.write(f"**Tipo:** {item['tipo_equipamento']}")
                c3.write(f"**Data:** {item['data_hora']}")
                if c4.button("🗑️", key=f"b_{idx}"): 
                    confirmar_exclusao_dialog(index=idx)
        
        st.divider()
        
        # Gera o arquivo Excel para as ações abaixo
        excel = exportar_para_excel(st.session_state['db_formularios'])
        
        if excel:
            col_down, col_mail = st.columns(2)
            
            # Ação 1: Download Direto
            col_down.subheader("Download")
            col_down.download_button(
                label="⬇️ Baixar Planilha (.xlsx)",
                data=excel,
                file_name=f"levantamento_{st.session_state['usuario_ativo']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            
            # Ação 2: Envio por E-mail
            col_mail.subheader("Enviar Relatório")
            email_dest = col_mail.text_input("E-mail do destinatário", placeholder="exemplo@empresa.com")
            if col_mail.button("📧 Enviar por E-mail", use_container_width=True):
                if email_dest:
                    with st.spinner("Enviando e-mail..."):
                        if enviar_email(excel, email_dest):
                            st.success(f"Relatório enviado para {email_dest}!")
                        else:
                            st.error("Erro ao enviar. Verifique sua conexão ou configurações de SMTP.")
                else:
                    st.warning("Por favor, insira um endereço de e-mail.")
        
        st.write("---")
        if st.button("⚠️ LIMPAR MEUS DADOS LOCALMENTE", use_container_width=True): 
            confirmar_exclusao_dialog(tipo="global")
elif menu == "3. Painel Admin":
    st.header("⚙️ Painel de Controle")
    t1, t2, t3 = st.tabs(["Técnicos", "Arquivos de Dados", "Planilha Base"])
    
    with t1:
        st.subheader("Gerenciar Acessos")
        with st.form("add_t"):
            n, s = st.text_input("Nome"), st.text_input("Senha", type="password")
            if st.form_submit_button("Cadastrar"):
                u_db = carregar_usuarios()
                u_db[n] = s
                salvar_usuarios(u_db)
                st.rerun()
        st.divider()
        u_db = carregar_usuarios()
        for u in u_db:
            if u != "Admin":
                col_n, col_d = st.columns([4, 1])
                col_n.write(f"👤 {u}")
                if col_d.button("Remover", key=f"rm_{u}"): excluir_usuario_dialog(u)

    with t2:
        st.subheader("📂 Recuperação e Exclusão de Arquivos")
        arquivos = sorted([f for f in os.listdir(".") if f.startswith("dados_") and f.endswith(".json")])
        if not arquivos:
            st.info("Nenhum arquivo de dados encontrado na pasta.")
        else:
            arq_sel = st.selectbox("Selecione um arquivo:", arquivos)
            dados_recup = carregar_dados_locais(path_especifico=arq_sel)
            
            st.write(f"**Registros encontrados:** {len(dados_recup)}")
            if dados_recup:
                df_view = pd.DataFrame([{"UC": d['cod_instalacao'], "Tipo": d['tipo_equipamento'], "Data": d['data_hora']} for d in dados_recup])
                st.table(df_view)
                
                c_rec_1, c_rec_2 = st.columns(2)
                excel_recup = exportar_para_excel(dados_recup)
                c_rec_1.download_button(f"⬇️ Exportar Excel", data=excel_recup, file_name=f"recuperado_{arq_sel}.xlsx", use_container_width=True)
                
                if c_rec_2.button("🔥 EXCLUIR ARQUIVO DO SERVIDOR", type="secondary", use_container_width=True):
                    excluir_arquivo_permanente_dialog(arq_sel)

    with t3:
        st.subheader("Modelo Mestre")
        arq_up = st.file_uploader("Novo Modelo", type=["xlsx"])
        if arq_up:
            with open(PLANILHA_PADRAO, "wb") as f: f.write(arq_up.getbuffer())
            st.success("Planilha base atualizada!")
