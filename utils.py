import streamlit as st
import pandas as pd
import io
import json
import os
import smtplib
import zipfile
import shutil
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import load_workbook
from datetime import datetime, timedelta, timezone
from io import BytesIO
from openpyxl import load_workbook
import io


def exportar_para_excel(lista_registros):
    """
    Usa o arquivo Levantamento_Base.xlsx como template e insere os dados
    preservando a formatação original.
    """
    if not lista_registros:
        return None

    # 1. Caminho do seu modelo original
    template_path = "Levantamento_Base.xlsx"
    
    if not os.path.exists(template_path):
        st.error("Arquivo template 'Levantamento_Base.xlsx' não encontrado!")
        return None

    # 2. Carregar o workbook (modelo)
    wb = load_workbook(template_path)
    
    # 3. Mapear os dados para as abas
    for reg in lista_registros:
        tipo = reg.get('tipo_equipamento')
        if tipo in wb.sheetnames:
            ws = wb[tipo]
            
            # Identificar os cabeçalhos da planilha para saber em qual coluna colar cada dado
            headers = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}
            
            # Preparar a linha de dados
            proxima_fila = ws.max_row + 1
            
            # Preencher dados do formulário (incluindo Localização que agora está em 'dados')
            dados_tecnicos = reg.get("dados", {})
            
            # Fallback para UC caso venha do legado
            if "Nome da Unidade Consumidora" not in dados_tecnicos and "cod_instalacao" in reg:
                dados_tecnicos["Nome da Unidade Consumidora"] = reg["cod_instalacao"]
                
            for campo, valor in dados_tecnicos.items():
                if campo in headers:
                    ws.cell(row=proxima_fila, column=headers[campo], value=valor)
            
            # Adicionar nomes das fotos na última coluna ou onde houver campo 'Fotos'
            if "Fotos" in headers:
                fotos = reg.get("fotos", [])
                nomes_fotos = ", ".join([f['nome_exportacao'] for f in fotos])
                ws.cell(row=proxima_fila, column=headers["Fotos"], value=nomes_fotos)

    # 4. Salvar o resultado em um buffer
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Constantes
PLANILHA_PADRAO_ADMIN = "Levantamento_Base.xlsx"
PASTA_FOTOS = "fotos_uploads"

# Garantir que a pasta de fotos existe
if not os.path.exists(PASTA_FOTOS):
    os.makedirs(PASTA_FOTOS)

# DATAS E CAMINHOS
def get_data_hora_br():
    fuso_br = timezone(timedelta(hours=-3))
    return datetime.now(fuso_br)

def get_user_data_path(nome_usuario=None):
    user = nome_usuario if nome_usuario else st.session_state.get('usuario_ativo')
    if user:
        nome_limpo = "".join(filter(str.isalnum, user))
        return f"dados_{nome_limpo}.json"
    return None

def get_user_template_path(nome_usuario=None):
    user = nome_usuario if nome_usuario else st.session_state.get('usuario_ativo')
    if user:
        nome_limpo = "".join(filter(str.isalnum, user))
        return f"template_{nome_limpo}.xlsx"
    return None

# PERSISTÊNCIA JSON
def salvar_dados_locais(dados):
    path = get_user_data_path()
    if path:
        with open(path, "w") as f: json.dump(dados, f)

def carregar_dados_locais(path_especifico=None):
    path = path_especifico if path_especifico else get_user_data_path()
    if path and os.path.exists(path):
        with open(path, "r") as f: return json.load(f)
    return []

# LÓGICA DE FOTOS

def salvar_fotos_local(lista_fotos_obj, cod_instalacao):
    """
    Recebe uma lista de dicionários: [{'arquivo': buffer, 'nome': 'descricao'}]
    Salva no disco e retorna os metadados.
    """
    caminhos_salvos = []
    
    # Garantir pasta
    if not os.path.exists(PASTA_FOTOS):
        os.makedirs(PASTA_FOTOS)
    
    for item in lista_fotos_obj:
        arquivo = item['arquivo']
        nome_personalizado = item['nome']
        
        if arquivo:
            # Tenta pegar extensão do arquivo original, se não tiver (câmera), usa .jpg
            nome_orig = getattr(arquivo, "name", "foto_camera.jpg")
            ext = os.path.splitext(nome_orig)[1]
            if not ext: ext = ".jpg"
            
            # Limpar nome definido pelo usuário para ser seguro no Windows/Linux
            nome_limpo = "".join(x for x in nome_personalizado if x.isalnum() or x in " -_")
            if not nome_limpo: nome_limpo = "imagem_sem_nome"
            
            # Timestamp para evitar sobrescrever
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S_%f")
            
            # Nome final físico: UC_TIMESTAMP_DESCRIÇÃO.ext
            nome_arquivo_final = f"{cod_instalacao}_{timestamp}_{nome_limpo}{ext}"
            caminho_completo = os.path.join(PASTA_FOTOS, nome_arquivo_final)
            
            # Salvar bytes
            with open(caminho_completo, "wb") as f:
                f.write(arquivo.getbuffer())
            
            caminhos_salvos.append({
                "caminho_fisico": caminho_completo,
                "nome_exportacao": f"{nome_limpo}{ext}", # Nome bonito para o ZIP
                "nome_original": nome_orig
            })
            
    return caminhos_salvos

# LÓGICA EXCEL E ZIP
def analisar_modelo_excel(file_content):
    try:
        buffer = io.BytesIO(file_content) if isinstance(file_content, bytes) else io.BytesIO(file_content.getvalue())
        wb = load_workbook(buffer, data_only=True)
        estrutura = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append({"nome": str(cell.value), "col_letter": cell.column_letter, "tipo": "texto", "opcoes": []})
            
            # Melhoria na detecção de validação de dados
            for dv in sheet.data_validations.dataValidation:
                if dv.type == "list":
                    for ref in str(dv.sqref).split():
                        col_letter = "".join(filter(str.isalpha, ref.split(':')[0]))
                        for h in headers:
                            if h["col_letter"] == col_letter:
                                h["tipo"] = "selecao"
                                formula = dv.formula1
                                # Ajuste: remove aspas e espaços extras para garantir que a lista seja lida
                                if formula:
                                    opcoes_limpas = formula.replace('"', '').split(',')
                                    h["opcoes"] = [op.strip() for op in opcoes_limpas]
                                    
            estrutura[sheet_name] = [{"nome": h["nome"], "tipo": h["tipo"], "opcoes": h["opcoes"]} for h in headers]
        return estrutura
    except Exception as e:
        st.error(f"Erro ao analisar o Excel: {e}")
        return {}

def carregar_modelo_atual():
    path_pessoal = get_user_template_path()
    if path_pessoal and os.path.exists(path_pessoal):
        with open(path_pessoal, "rb") as f:
            content = f.read()
            st.session_state['planilha_modelo'] = io.BytesIO(content)
            st.session_state['estrutura_modelo'] = analisar_modelo_excel(content)
            st.session_state['origem_modelo'] = "Pessoal"
    elif os.path.exists(PLANILHA_PADRAO_ADMIN):
        with open(PLANILHA_PADRAO_ADMIN, "rb") as f:
            content = f.read()
            st.session_state['planilha_modelo'] = io.BytesIO(content)
            st.session_state['estrutura_modelo'] = analisar_modelo_excel(content)
            st.session_state['origem_modelo'] = "Padrão do Sistema"

def gerar_zip_exportacao(dados_lista):
    """
    Gera um arquivo ZIP contendo o Excel de levantamento e uma pasta com as fotos.
    """
    if 'planilha_modelo' not in st.session_state: return None
    
    # 1. Gerar o Excel em memória
    st.session_state['planilha_modelo'].seek(0)
    book = load_workbook(st.session_state['planilha_modelo'])
    
    for registro in dados_lista:
        if registro['tipo_equipamento'] in book.sheetnames:
            sheet = book[registro['tipo_equipamento']]
            
            # Mapeamento dinâmico baseado no cabeçalho
            headers = {sheet.cell(row=1, column=i).value: i for i in range(1, sheet.max_column + 1)}
            
            nova_linha_dados = registro['dados']
            
            # Preparar array para append (garantindo ordem)
            # Como append adiciona no fim, precisamos garantir que as colunas batem com os valores
            # Melhor abordagem com openpyxl para dados esparsos: escrever celula a celula na nova linha
            next_row = sheet.max_row + 1
            
            for k, v in nova_linha_dados.items():
                if k in headers:
                    sheet.cell(row=next_row, column=headers[k], value=v)
            
            # Tratamento de Fotos no Excel (se houver coluna)
            if "Fotos" in headers and "fotos" in registro:
                 nomes_fotos = ", ".join([f['nome_exportacao'] for f in registro["fotos"]])
                 sheet.cell(row=next_row, column=headers["Fotos"], value=nomes_fotos)
    
    excel_buffer = io.BytesIO()
    book.save(excel_buffer)
    excel_buffer.seek(0)
    
    # 2. Criar o ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Adicionar Excel
        zf.writestr("Levantamento_Cargas.xlsx", excel_buffer.getvalue())
        
        # Adicionar Fotos
        for registro in dados_lista:
            if "fotos" in registro and registro["fotos"]:
                uc = registro.get("cod_instalacao", "SemUC")
                tipo = registro.get("tipo_equipamento", "Geral")
                
                # Pasta dentro do ZIP para organizar
                folder_path = f"Fotos/{uc} - {tipo}/"
                
                for foto in registro["fotos"]:
                    caminho_origem = foto["caminho_fisico"]
                    if os.path.exists(caminho_origem):
                        # Nome dentro do ZIP
                        zf.write(caminho_origem, arcname=f"{folder_path}{foto['nome_exportacao']}")
    
    zip_buffer.seek(0)
    return zip_buffer

# EMAIL (Atualizado para enviar ZIP se tiver fotos ou apenas Excel)
def enviar_email(arquivo_buffer, destinatario, is_zip=False):
    try:
        msg = MIMEMultipart()
        msg['From'] = "levantamento.poupenergia@gmail.com"
        msg['To'] = destinatario
        msg['Subject'] = f"Levantamento {st.session_state['usuario_ativo']} - {get_data_hora_br().strftime('%d/%m/%Y')}"
        msg.attach(MIMEText("Segue em anexo o levantamento realizado.", 'plain'))
        
        part = MIMEBase('application', 'zip' if is_zip else 'octet-stream')
        part.set_payload(arquivo_buffer.getvalue())
        encoders.encode_base64(part)
        
        filename = "levantamento_completo.zip" if is_zip else "levantamento.xlsx"
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        
        msg.attach(part)
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        # Nota: Idealmente usar variáveis de ambiente para senhas
        server.login("levantamento.poupenergia@gmail.com", "kiqplowxqprcugjc")
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(e)
        return False
